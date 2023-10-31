import openpyxl

workbook1 = openpyxl.Workbook()

sheet1 = workbook1.active

sheet1['A1'] = 'Имя'
sheet1['B1'] = 'Электронная почта'

users = [
    ('Иван', 'ivan@example.com'),
    ('Мария', 'maria@example.com'),
    ('Петр', 'petr@example.com'),
]

for index, user in enumerate(users, start=2):
    sheet1[f'A{index}'] = user[0]
    sheet1[f'B{index}'] = user[1]

workbook1.save('пользователи.xlsx')

workbook2 = openpyxl.load_workbook('пользователи.xlsx')

sheet2 = workbook2.active

for row in sheet2.iter_rows(min_row=2, values_only=True):
    name, email = row
    print(f'Имя: {name}, Электронная почта: {email}')

workbook2.close()

######################################################

import requests

url = 'https://jsonplaceholder.typicode.com/users'

response = requests.get(url)

if response.status_code == 200:
    users = response.json()

    for user in users:
        print(user['name'])
else:
    print(f'Ошибка при выполнении запроса. Код ответа: {response.status_code}')
